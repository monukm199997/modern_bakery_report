from io import BytesIO
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from app.dependencies.auth import get_current_user
from app.core.database import get_db
from app.reports.vehicles_permit_report.schemas.vehicles_schema import VehiclesPermitRequest
from app.reports.vehicles_permit_report.utils.vehicles_helper import prepare_dashboard_context
from app.common.apply_payload_permissions import apply_payload_permissions
from app.reports.vehicles_permit_report.utils.vehicles_sql_query import SELECT_QUERY, JOIN_QUERY

router = APIRouter(tags=["Vehicles Permit Report"], dependencies=[Depends(get_current_user)])

HEADER_FILL = PatternFill(
    start_color="903442",
    end_color="903442",
    fill_type="solid"
)

HEADER_FONT = Font(
    bold=True,
    color="FFFFFF"
)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

CENTER = Alignment(horizontal="center", vertical="center")


@router.post("/vehicles-permit-export")
def vehicles_permit_export(
    payload: VehiclesPermitRequest,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    payload = apply_payload_permissions(payload, db, current_user)
    ctx = prepare_dashboard_context(payload)

    query = f"""
        SELECT
           {SELECT_QUERY}
        {JOIN_QUERY}
        WHERE {ctx['where_sql']}
        ORDER BY vp.expiry_date
    """

    rows = db.execute(text(query), ctx["params"]).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Vehicles Permit"

    headers = [
        "Vehicle Number Plate",
        "Region",
        "Permit Number",
        "Permit Expiry Date",
        "Registration Card Number",
        "Registration Card Expiry Date"
    ]

    ws.append(headers)

    # Header Style
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = CENTER

    # Data
    for row in rows:
        ws.append([
            row.vehicles_number_plate,
            row.region,
            row.permit_number,
            row.permit_expiry_date.strftime("%d-%m-%Y") if row.permit_expiry_date else "",
            row.registration_card_number,
            row.registration_card_expiry_date.strftime("%d-%m-%Y") if row.registration_card_expiry_date else ""
        ])

    # Apply borders
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = THIN_BORDER

    # Auto-fit columns
    for column in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = min(max_length + 4, 40)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="Vehicles_Permit_Report.xlsx"'
        },
    )