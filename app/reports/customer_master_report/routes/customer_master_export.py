from io import BytesIO
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from app.dependencies.auth import get_current_user
from app.core.database import get_db
from app.common.apply_payload_permissions import apply_payload_permissions
from app.reports.customer_master_report.schemas.customer_master_schema import CustomerMasterRequest
from app.reports.customer_master_report.utils.customer_master_helper import prepare_dashboard_context
from app.reports.customer_master_report.utils.customer_master_sql_query import SELECT_QUERY, JOIN_QUERY

router = APIRouter(tags=["Customer Master Report"], dependencies=[Depends(get_current_user)])

# Styles
HEADER_FILL = PatternFill(
    start_color="903442",
    end_color="903442",
    fill_type="solid"
)

HEADER_FONT = Font(
    bold=True,
    color="FFFFFF"
)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

CENTER = Alignment(horizontal="center", vertical="center")


@router.post("/customer-master-export")
def customer_master_export(
    payload: CustomerMasterRequest,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    payload = apply_payload_permissions(payload, db, current_user)
    ctx = prepare_dashboard_context(payload)

    query = f"""
        SELECT
            {SELECT_QUERY}
        {JOIN_QUERY}
        WHERE {ctx['where_sql']}
        ORDER BY ac.dateof_creation
    """

    rows = db.execute(text(query), ctx["params"]).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Customer Master"

    headers = [
        "Customer Code",
        "Customer Name",
        "Date of Creation",
        "Status",
        "Route Code",
        "Route Name",
        "Salesman Code",
        "Salesman Name",
        "Outlet Channel",
        "TL Number",
        "TIN No",
        "Customer Type",
        "Customer Group",
        "Payment Terms",
        "Address",
        "Region",
        "Latitude",
        "Longitude"
    ]

    ws.append(headers)

    # Header Style
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = CENTER

    # Data
    for row in rows:
        ws.append([
            row.customer_code,
            row.customer_name,
            row.dateof_creation.strftime("%d-%m-%Y") if row.dateof_creation else "",
            row.status,
            row.route_code,
            row.route_name,
            row.salesman_code,
            row.salesman_name,
            row.outlet_channel,
            row.tl_number,
            row.tin_no,
            row.customer_type,
            row.cust_group,
            row.payment_terms,
            row.address,
            row.region_name,
            row.latitude,
            row.longitude
        ])

    # Border for all data cells
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = THIN_BORDER

    # Auto-fit column widths
    for column in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = min(max_length + 4, 40)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="Customer_Master_Report.xlsx"'
        },
    )