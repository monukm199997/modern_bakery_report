import os
import tempfile
from datetime import datetime

from fastapi import APIRouter, Depends
from fastapi.responses import FileResponse
from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from starlette.background import BackgroundTask

from app.core.database import get_db
from app.dependencies.auth import get_current_user
from app.reports.target_commison_report.schemas.schemas import SalesAchievementSchema
# Data pipeline now lives in the table-view module; the export only renders it.
from app.reports.target_commison_report.routes.target_commison_table import (
    get_sales_achievement_data,
)


router = APIRouter(tags=["Target Commission Report"], dependencies=[Depends(get_current_user)])


YELLOW = PatternFill(start_color="993442", end_color="993442", fill_type="solid")
BLUE = PatternFill(start_color="CC6677", end_color="CC6677", fill_type="solid")
TOTAL = PatternFill(start_color="CC6677", end_color="CC6677", fill_type="solid")
THIN = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BOLD = Font(bold=True, size=10, color="FFFFFF")
CENTER = Alignment(horizontal="center", vertical="center")


REGION_SUB_HEADERS = [
    "Route", "Code", "Salesman", "Supervisor",
    "Sales", "Returns", "Retn %", "Net Sales", "Target", "Achv %",
    "Sales", "Returns", "Retn %", "Net Sales", "Target", "Achv %",
    "Achievement (Projected)", "Target", "Achv %",
]
TOTAL_COLS = len(REGION_SUB_HEADERS)  # 19

SUMMARY_SUB_HEADERS = [
    "Region",
    "Net Sales", "Target", "Achv %",     # DAILY
    "Net Sales", "Target", "Achv %",     # CUMULATIVE
    "Projected", "Target", "Achv %",     # MONTHLY
]


def _style_header(cell):
    cell.fill = YELLOW
    cell.font = BOLD
    cell.border = BORDER
    cell.alignment = CENTER


def _write_row(ws, row_idx, values, fill=None, bold=False):
    for col, value in enumerate(values, 1):
        cell = ws.cell(row_idx, col)
        # Skip cells that are part of a merge but not the top-left anchor
        if isinstance(cell, MergedCell):
            continue
        cell.value = value
        cell.border = BORDER
        cell.alignment = CENTER
        if fill:
            cell.fill = fill
        if bold:
            cell.font = BOLD
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            cell.number_format = "#,##0.00"


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        try:
            column_letter = get_column_letter(col[0].column)
        except AttributeError:
            continue
        for cell in col:
            if isinstance(cell, MergedCell):
                continue
            try:
                if cell.value is not None:
                    length = min(len(str(cell.value)), 25)
                    max_len = max(max_len, length)
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = max_len + 3


def build_workbook(grouped_data, from_date: str, to_date: str, group_label: str = "Region"):
    wb = Workbook()
    wb.remove(wb.active)

    summary_ws = wb.create_sheet("Summary", 0)

    # Group headers (row 1)
    summary_ws.cell(1, 1, "")  # Group label (Company/Region/Route) sits empty above
    _style_header(summary_ws.cell(1, 1))
    # DAILY group: cols 2-4
    summary_ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=4)
    summary_ws.cell(1, 2, f"DAILY ({to_date})")
    _style_header(summary_ws.cell(1, 2))
    # CUMULATIVE group: cols 5-7
    summary_ws.merge_cells(start_row=1, start_column=5, end_row=1, end_column=7)
    summary_ws.cell(1, 5, f"CUMULATIVE ({from_date} – {to_date})")
    _style_header(summary_ws.cell(1, 5))
    # MONTHLY group: cols 8-10
    summary_ws.merge_cells(start_row=1, start_column=8, end_row=1, end_column=10)
    summary_ws.cell(1, 8, "ESTIMATED MONTHLY SALES")
    _style_header(summary_ws.cell(1, 8))

    # Sub-headers (row 2) — first column label is dynamic ("Company", "Region", or "Route")
    for col, header in enumerate(SUMMARY_SUB_HEADERS, 1):
        label = group_label if col == 1 else header
        _style_header(summary_ws.cell(2, col, label))

    summary_row = 3
    grand = {
        "daily_net": 0, "daily_target": 0,
        "mtd_net": 0, "mtd_target": 0,
        "projected": 0, "monthly_target": 0,
    }

    # ------------------------------------------------------------
    # GROUP SHEETS (one tab per Company, Region, or Route — depending
    # on which filter/grouping is active)
    # ------------------------------------------------------------
    for group_name, rows in grouped_data.items():
        ws = wb.create_sheet((group_name or "Unknown")[:30])

        # Group headers (row 1)
        # Cols A-D: blank header band (sits above Route + Code + Salesman + Supervisor)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
        _style_header(ws.cell(1, 1, ""))
        # DAILY (cols E-J = 5..10)
        ws.merge_cells(start_row=1, start_column=5, end_row=1, end_column=10)
        ws.cell(1, 5, f"DAILY ({to_date})")
        _style_header(ws.cell(1, 5))
        # CUMULATIVE (cols K-P = 11..16)
        ws.merge_cells(start_row=1, start_column=11, end_row=1, end_column=16)
        ws.cell(1, 11, f"CUMULATIVE ({from_date} – {to_date})")
        _style_header(ws.cell(1, 11))
        # ESTIMATED MONTHLY SALES (cols Q-S = 17..19)
        ws.merge_cells(start_row=1, start_column=17, end_row=1, end_column=19)
        ws.cell(1, 17, "ESTIMATED MONTHLY SALES")
        _style_header(ws.cell(1, 17))

        # Sub-headers (row 2)
        for col, header in enumerate(REGION_SUB_HEADERS, 1):
            _style_header(ws.cell(2, col, header))

        # Freeze header rows 1-2 so they stay visible while scrolling
        ws.freeze_panes = "A3"

        row_idx = 3
        tot = {
            "daily_sales": 0, "daily_returns": 0, "daily_net": 0,
            "daily_target": 0,
            "mtd_sales": 0, "mtd_returns": 0, "mtd_net": 0,
            "mtd_target": 0,
            "projected": 0, "monthly_target": 0,
        }

        for item in rows:
            # Order MUST match REGION_SUB_HEADERS
            _write_row(ws, row_idx, [
                item["route_code"],
                item["osa_code"],
                item["salesman"],
                item["supervisor"],
                # DAILY block
                item["daily_sales"],
                item["daily_returns"],
                round(item["daily_ret_pct"], 2),
                item["daily_net"],
                round(item["daily_target"], 2),
                round(item["daily_ach"], 2),
                # CUMULATIVE block
                item["cumulative_sales"],
                item["cumulative_returns"],
                round(item["mtd_ret_pct"], 2),
                item["cumulative_net"],
                round(item["mtd_target"], 2),
                round(item["cumulative_ach"], 2),
                # MONTHLY block
                round(item["projected_sales"], 2),
                item["monthly_target"],
                round(item["projected_ach"], 2),
            ])
            tot["daily_sales"] += item["daily_sales"]
            tot["daily_returns"] += item["daily_returns"]
            tot["daily_net"] += item["daily_net"]
            tot["daily_target"] += item["daily_target"]
            tot["mtd_sales"] += item["cumulative_sales"]
            tot["mtd_returns"] += item["cumulative_returns"]
            tot["mtd_net"] += item["cumulative_net"]
            tot["mtd_target"] += item["mtd_target"]
            tot["projected"] += item["projected_sales"]
            tot["monthly_target"] += item["monthly_target"]
            row_idx += 1

        # ----- Region salesman total -----
        daily_ret_pct = (
            tot["daily_returns"] / tot["daily_sales"] * 100
            if tot["daily_sales"] else 0
        )
        mtd_ret_pct = (
            tot["mtd_returns"] / tot["mtd_sales"] * 100
            if tot["mtd_sales"] else 0
        )
        daily_pct = (
            tot["daily_net"] / tot["daily_target"] * 100
            if tot["daily_target"] else 0
        )
        mtd_pct = (
            tot["mtd_net"] / tot["mtd_target"] * 100
            if tot["mtd_target"] else 0
        )
        month_pct = (
            tot["projected"] / tot["monthly_target"] * 100
            if tot["monthly_target"] else 0
        )

        _write_row(ws, row_idx, [
            f"{group_name} TOTAL", "", "", "",
            tot["daily_sales"],
            tot["daily_returns"],
            round(daily_ret_pct, 2),
            tot["daily_net"],
            round(tot["daily_target"], 2),
            round(daily_pct, 2),
            tot["mtd_sales"],
            tot["mtd_returns"],
            round(mtd_ret_pct, 2),
            tot["mtd_net"],
            round(tot["mtd_target"], 2),
            round(mtd_pct, 2),
            round(tot["projected"], 2),
            tot["monthly_target"],
            round(month_pct, 2),
        ], fill=BLUE, bold=True)
        ws.merge_cells(start_row=row_idx, start_column=1,
                       end_row=row_idx, end_column=4)

        # ----- Summary sheet row for this region -----
        _write_row(summary_ws, summary_row, [
            group_name,
            tot["daily_net"],
            round(tot["daily_target"], 2),
            round(daily_pct, 2),
            tot["mtd_net"],
            round(tot["mtd_target"], 2),
            round(mtd_pct, 2),
            round(tot["projected"], 2),
            tot["monthly_target"],
            round(month_pct, 2),
        ])
        summary_row += 1

        grand["daily_net"] += tot["daily_net"]
        grand["daily_target"] += tot["daily_target"]
        grand["mtd_net"] += tot["mtd_net"]
        grand["mtd_target"] += tot["mtd_target"]
        grand["projected"] += tot["projected"]
        grand["monthly_target"] += tot["monthly_target"]

        _auto_width(ws)

    # ------------------------------------------------------------
    # GRAND TOTAL on summary
    # ------------------------------------------------------------
    g_daily_pct = (
        grand["daily_net"] / grand["daily_target"] * 100
        if grand["daily_target"] else 0
    )
    g_mtd_pct = (
        grand["mtd_net"] / grand["mtd_target"] * 100
        if grand["mtd_target"] else 0
    )
    g_month_pct = (
        grand["projected"] / grand["monthly_target"] * 100
        if grand["monthly_target"] else 0
    )

    _write_row(summary_ws, summary_row, [
        "GRAND TOTAL",
        grand["daily_net"],
        round(grand["daily_target"], 2),
        round(g_daily_pct, 2),
        grand["mtd_net"],
        round(grand["mtd_target"], 2),
        round(g_mtd_pct, 2),
        round(grand["projected"], 2),
        grand["monthly_target"],
        round(g_month_pct, 2),
    ], fill=TOTAL, bold=True)

    _auto_width(summary_ws)
    return wb


def _cleanup(path: str):
    try:
        os.remove(path)
    except OSError:
        pass


@router.post("/sales-achievement-export")
def sales_achievement_export(
    payload: SalesAchievementSchema,
    db: Session = Depends(get_db),
):
    # Same data the table view builds — export only renders it to Excel.
    grouped_data, meta = get_sales_achievement_data(db, payload)

    group_by = meta["group_by"]
    group_label = {"company": "Company", "region": "Region", "route_code": "Route"}[group_by]
    file_prefix = f"sales_achievement_by_{group_by}_"

    wb = build_workbook(
        grouped_data,
        from_date=meta["from_date"],
        to_date=meta["to_date"],
        group_label=group_label,
    )

    # Write to a temp file and let FastAPI delete it after sending.
    fd, tmp_path = tempfile.mkstemp(suffix=".xlsx", prefix=file_prefix)
    os.close(fd)
    wb.save(tmp_path)

    download_name = (
        f"{file_prefix}{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    )

    return FileResponse(
        path=tmp_path,
        filename=download_name,
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        background=BackgroundTask(_cleanup, tmp_path),
    )