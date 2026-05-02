from app.reports.sales_report.schemas.schemas import SalesReportRequest
from app.utils.helper import validate_mandatory, choose_granularity, quantity_expr_sql
from openpyxl.styles import Font, PatternFill


def build_query_parts(payload: SalesReportRequest):
    joins = []
    where_fragments = []
    params = {}

    where_fragments.append("ih.invoice_date BETWEEN :from_date AND :to_date")
    params["from_date"] = payload.from_date
    params["to_date"] = payload.to_date

    if payload.display_quantity and payload.display_quantity.lower() == "without_free_good":
        where_fragments.append("id.item_total <> 0")
    
    if payload.company_ids == []:
        where_fragments.append("1 = 0")
    elif payload.company_ids:
        joins.append("LEFT JOIN tbl_route rt ON rt.id = ih.route_id")
        where_fragments.append("ih.company_id = ANY(:company_ids)")
        params["company_ids"] = payload.company_ids

    if payload.region_ids == []:
        where_fragments.append("1 = 0")
    elif payload.region_ids:
        joins.append("LEFT JOIN tbl_route rt ON rt.id = ih.route_id")
        where_fragments.append("rt.region_id = ANY(:region_ids)")
        params["region_ids"] = payload.region_ids

    if payload.route_ids == []:
        where_fragments.append("1 = 0")
    elif payload.route_ids:
        where_fragments.append("ih.route_id = ANY(:route_ids)")
        params["route_ids"] = payload.route_ids
        
    if payload.salesman_ids == []:
        where_fragments.append("1 = 0")
    elif payload.salesman_ids:
        where_fragments.append("ih.salesman_id = ANY(:salesman_ids)")
        params["salesman_ids"] = payload.salesman_ids

    if payload.item_category_ids == []:
        where_fragments.append("1 = 0")
    elif payload.item_category_ids:
        where_fragments.append("it.category_id = ANY(:item_category_ids)")
        params["item_category_ids"] = payload.item_category_ids

    if payload.item_ids == []:
        where_fragments.append("1 = 0")
    elif payload.item_ids:
        where_fragments.append("id.item_id = ANY(:item_ids)")
        params["item_ids"] = payload.item_ids

    if payload.customer_channel_ids == []:
        where_fragments.append("1 = 0")
    elif payload.customer_channel_ids:
        where_fragments.append("ac.outlet_channel_id = ANY(:customer_channel_ids)")
        params["customer_channel_ids"] = payload.customer_channel_ids

    joins = list(dict.fromkeys(joins))
    return joins, where_fragments, params




def prepare_dashboard_context(payload: SalesReportRequest):
    validate_mandatory(payload)

    granularity, period_label_sql, order_by_sql = choose_granularity(
        payload.from_date, payload.to_date
    )

    joins, where_fragments, params = build_query_parts(payload)
    where_sql = " AND ".join(where_fragments)
    join_sql = "\n".join(joins)

    quantity = quantity_expr_sql()
    value_expr = (
        quantity if payload.search_type.lower() == "quantity"
        else "SUM(id.item_total)"
    )

    return {
        "granularity": granularity,
        "period_label_sql": period_label_sql,
        "order_by_sql": order_by_sql,
        "join_sql": join_sql,
        "where_sql": where_sql,
        "params": params,
        "value_expr": value_expr,
    }


def choose_export_granularity(
    from_date_str: str,
    to_date_str: str,
    download_type: str
):
    if download_type == "daily":
        return (
            "daily",
            "TO_CHAR(ih.invoice_date, 'YYYY-MM-DD')",
            "ih.invoice_date"
        )

    elif download_type == "weekly":
        return (
            "weekly",
            f"""
            CONCAT(
                TO_CHAR(
                    GREATEST(
                        DATE_TRUNC('week', ih.invoice_date),
                        DATE '{from_date_str}'
                    ),
                    'DD Mon'
                ),
                ' - ',
                TO_CHAR(
                    LEAST(
                        DATE_TRUNC('week', ih.invoice_date) + INTERVAL '6 days',
                        DATE '{to_date_str}'
                    ),
                    'DD Mon'
                )
            )
            """,
            "DATE_TRUNC('week', ih.invoice_date)"
        )

    elif download_type == "monthly":
        return (
            "monthly",
            "TO_CHAR(DATE_TRUNC('month', ih.invoice_date), 'Mon-YYYY')",
            "DATE_TRUNC('month', ih.invoice_date)"
        )

    elif download_type == "yearly":
        return (
            "yearly",
            "TO_CHAR(DATE_TRUNC('year', ih.invoice_date), 'YYYY')",
            "DATE_TRUNC('year', ih.invoice_date)"
        )

    # default behavior
    return choose_granularity(from_date_str, to_date_str)

def style_sheet(ws):

        # Freeze pane
        ws.freeze_panes = "D2"

        # Header style
        header_fill = PatternFill(
            start_color="993442",
            end_color="993442",
            fill_type="solid"
        )

        header_font = Font(
            bold=True,
            color="FFFFFF"
        )

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        green_font = Font(
            bold=True,
            color="008000"
        )

        red_font = Font(
            bold=True,
            color="FF0000"
        )

        for row in ws.iter_rows(min_row=2):

            item_code = row[0].value
            item_name = (
                str(row[1].value).strip()
                if row[1].value is not None
                else ""
            )

            # Total row
            if item_name.lower() == "total":
                for cell in row:
                    cell.font = red_font

            # Category row
            elif item_code in ("", None):
                for cell in row:
                    cell.font = green_font

        # Auto width
        for column_cells in ws.columns:
            max_len = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells
            )
            ws.column_dimensions[
                column_cells[0].column_letter
            ].width = min(max_len + 3, 40)



def get_level_config(payload):
    if payload.item_ids:
        return {
            "level_id_col": "id.item_id",
            "level_name_col": "it.name",
            "level_label": "item_name",
            "level_join": ""
        }

    elif payload.item_category_ids:
        return {
            "level_id_col": "it.category_id",
            "level_name_col": "cat.category_name",
            "level_label": "item_category",
            "level_join": ""
        }

    elif payload.customer_channel_ids:
        return {
            "level_id_col": "ac.outlet_channel_id",
            "level_name_col": "ch.outlet_channel",
            "level_label": "channel_name",
            "level_join": """
                LEFT JOIN outlet_channel ch ON ch.id = ac.outlet_channel_id
            """
        }

    elif payload.salesman_ids:
        return {
            "level_id_col": "ih.salesman_id",
            "level_name_col": "sm.name",
            "level_label": "salesman_name",
            "level_join": "LEFT JOIN salesman sm ON sm.id = ih.salesman_id"
        }

    elif payload.route_ids:
        return {
            "level_id_col": "ih.route_id",
            "level_name_col": "rt.route_name",
            "level_label": "route_name",
            "level_join": ""
        }

    elif payload.region_ids:
        return {
            "level_id_col": "rt.region_id",
            "level_name_col": "r.region_name",
            "level_label": "region_name",
            "level_join": """
                LEFT JOIN tbl_region r ON r.id = rt.region_id
            """
        }

    elif payload.company_ids:
        return {
            "level_id_col": "ih.company_id",
            "level_name_col": "c.company_name",
            "level_label": "company_name",
            "level_join": "LEFT JOIN tbl_company c ON c.id = ih.company_id"
        }

    return {
        "level_id_col": "ih.company_id",
        "level_name_col": "c.company_name",
        "level_label": "company_name",
        "level_join": "LEFT JOIN tbl_company c ON c.id = ih.company_id"
    }