from io import BytesIO
from datetime import datetime
import pandas as pd
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy import text
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from app.core.database import get_db
import numbers
from app.utils.constant import MONTH_ORDER
from app.reports.sales_report.utils.sql_query_helper import SELECT,GROUP_BY,CHANNEL_JOIN_SQL,ITEM_JOIN_SQL
from app.reports.customer_sales_report.utils.sql_query_helper import BASE_SQL
from app.dependencies.auth import get_current_user
from app.reports.sales_report.schemas.schemas import (
    SalesReportRequest
)
from app.reports.sales_report.utils.sales_report_helper import (
    prepare_dashboard_context,
)

router = APIRouter(tags=["Sales Report"],dependencies=[Depends(get_current_user)])

@router.post('/group-sales-matrix-export')
def group_sales_matrix_export(payload: SalesReportRequest, db: Session = Depends(get_db)):
    ctx = prepare_dashboard_context(payload)
    current_year = datetime.now().year
    query = f"""
        WITH yearly_sales AS (
            SELECT
                {SELECT}
                EXTRACT(YEAR FROM ih.invoice_date)::text AS period,
                {ctx['value_expr']} AS value,
                'yearly' AS period_type,
                EXTRACT(YEAR FROM ih.invoice_date) AS sort_order
            {BASE_SQL}
            {CHANNEL_JOIN_SQL}
            {ITEM_JOIN_SQL}
            {ctx['join_sql']}
            WHERE
                {ctx['where_sql']}
                AND EXTRACT(YEAR FROM ih.invoice_date)
                    < :current_year

            GROUP BY
                {GROUP_BY}
                EXTRACT(YEAR FROM ih.invoice_date)
        ),
        monthly_sales AS (
            SELECT
                {SELECT}
                TO_CHAR(
                    ih.invoice_date,
                    'Mon-YY'
                ) AS period,
                {ctx['value_expr']} AS value,
                'monthly' AS period_type,
                EXTRACT(MONTH FROM ih.invoice_date) + 100 AS sort_order
            {BASE_SQL}
            {CHANNEL_JOIN_SQL}
            {ITEM_JOIN_SQL}
            {ctx['join_sql']}
            WHERE
                {ctx['where_sql']}
                AND EXTRACT(YEAR FROM ih.invoice_date)
                    = :current_year
            GROUP BY
                {GROUP_BY}
                TO_CHAR(
                    ih.invoice_date,
                    'Mon-YY'
                ),
                EXTRACT(MONTH FROM ih.invoice_date)
        )
        SELECT *
        FROM yearly_sales
        UNION ALL
        SELECT *
        FROM monthly_sales
        ORDER BY
            "Item",
            sort_order
    """
    params = {**ctx['params'], 'current_year': current_year}
    rows = db.execute(text(query),params).fetchall()
    data = [dict(r._mapping) for r in rows]

    if not data:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Sales Matrix'
        ws['A1'] = 'No Data Found'

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition':
                'attachment; filename=sales_matrix.xlsx'
            }
        )
    
    df = pd.DataFrame(data)
    
    pivot_df = df.pivot_table(
        index=['Product','Item','PK','Unit'],
        columns='period',
        values='value',
        aggfunc='sum',
        fill_value=0
    )
    pivot_df.reset_index(inplace=True)

    yearly_cols = sorted([
        col for col in pivot_df.columns
        if str(col).isdigit()
    ])

    month_order = MONTH_ORDER

    monthly_cols = []
    for month in month_order:
        for col in pivot_df.columns:
            if str(col).startswith(month):
                monthly_cols.append(col)

    final_columns = [
       'Product','Item','PK','Unit'
    ] + yearly_cols + monthly_cols

    pivot_df = pivot_df[final_columns]

    wb = Workbook()
    ws = wb.active
    ws.title = 'Sales Matrix'

    header_fill = PatternFill(
        start_color='993442',
        end_color='993442',
        fill_type='solid'
    )

    for col_num, column_name in enumerate(final_columns, 1):
        cell = ws.cell(
            row=1,
            column=col_num,
            value=str(column_name)
        )
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal='center'
        )

    for row_num, row_data in enumerate(
        pivot_df.values,
        2
    ):

        for col_num, value in enumerate(row_data, 1):
            ws.cell(
                row=row_num,
                column=col_num,
                value=float(value)
                if isinstance(value, (int, float))
                else value
            )

    for column_cells in ws.columns:
        length = max(
            len(str(cell.value))
            if cell.value else 0
            for cell in column_cells
        )
        column_letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[
            column_letter
        ].width = length + 5

    total_row = ws.max_row + 1

    total_fill = PatternFill(
        start_color='D9EAD3',
        end_color='D9EAD3',
        fill_type='solid'
    )
    # TOTAL label
    total_cell = ws.cell(
        row=total_row,
        column=1,
        value='TOTAL'
    )
    total_cell.font = Font(bold=True)
    total_cell.fill = total_fill
    numeric_start_col = 5

    for col in range(numeric_start_col, ws.max_column + 1):

        total = 0

        # Data starts from row 2
        for row in range(2, total_row):

            cell_value = ws.cell(
                row=row,
                column=col
            ).value

            if isinstance(cell_value, numbers.Number):
                total += float(cell_value)

        total_value_cell = ws.cell(
            row=total_row,
            column=col,
            value=round(total, 2)
        )

        total_value_cell.font = Font(bold=True)
        total_value_cell.fill = total_fill

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = (
        f'sales_matrix_'
        f'{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition':
            f'attachment; filename={filename}'
        }
    )