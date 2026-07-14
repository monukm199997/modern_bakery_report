from fastapi import APIRouter, Depends
from sqlalchemy.orm import Session
from sqlalchemy import text
from fastapi.responses import FileResponse

import openpyxl
from tempfile import NamedTemporaryFile
from openpyxl.styles import (
    PatternFill,
    Font,
    Border,
    Side,
    Alignment,
)

from app.core.database import get_db
from app.dependencies.auth import get_current_user
from app.reports.inventory_movement_report.schemas.inventory_movement_schema import (
    InventoryMovementRequest,
)
from app.reports.inventory_movement_report.utils.inventory_movement_helper import (
    prepare_inventory_movement_context,
)
from app.reports.inventory_movement_report.routes.inventory_movement_tableview import (
    build_inventory_movement_query,
)

router = APIRouter(
    tags=["Inventory Movement Report"], dependencies=[Depends(get_current_user)]
)


HEADER_FILL = PatternFill(
    start_color="903442",
    end_color="903442",
    fill_type="solid",
)

HEADER_FONT = Font(
    color="FFFFFF",
    bold=True,
)

TOTAL_FONT = Font(
    color="FFFFFF",
    bold=True,
)

BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")


def pretty_header(header: str) -> str:
    return header.replace("_", " ").title()


def to_float(value):
    try:
        return float(value or 0)
    except Exception:
        return 0


@router.post("/inventory-movement-export")
def inventory_movement_export(
    payload: InventoryMovementRequest,
    db: Session = Depends(get_db),
):
    ctx = prepare_inventory_movement_context(payload)

    query = build_inventory_movement_query(ctx)

    rows = db.execute(text(query), ctx["params"]).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory Movement Report"

    if not rows:
        ws.append(["No Data Found"])

        cell = ws["A1"]
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = CENTER

    else:
        headers = list(rows[0]._mapping.keys())

        if "salesman_id" in headers:
            headers.remove("salesman_id")

        ws.append([pretty_header(h) for h in headers])

        # Header Style
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = BORDER
            cell.alignment = CENTER

        # Data
        for row in rows:
            mapping = row._mapping
            ws.append([mapping.get(header) for header in headers])

        numeric_headers = {
            "open_stock",
            "load_qty",
            "sold_qty",
            "grv_qty",
            "van_return_qty",
            "close_stock",
        }

        # Body Style
        for excel_row in ws.iter_rows(min_row=2):
            for cell in excel_row:
                cell.border = BORDER

                header = headers[cell.column - 1]

                if header in numeric_headers:
                    cell.alignment = RIGHT
                    cell.number_format = "#,##0.000"

                    if isinstance(cell.value, (int, float)) and cell.value < 0:
                        cell.font = Font(color="FF0000")

        # Total Row
        total_row = ws.max_row + 1

        for col_idx, header in enumerate(headers, start=1):

            cell = ws.cell(row=total_row, column=col_idx)

            cell.fill = HEADER_FILL
            cell.font = TOTAL_FONT
            cell.border = BORDER

            if col_idx == 1:
                cell.value = "Total"
                cell.alignment = CENTER
                continue

            if header in numeric_headers:

                total = sum(to_float(r._mapping.get(header)) for r in rows)

                cell.value = total
                cell.alignment = RIGHT
                cell.number_format = "#,##0.000"

                if total < 0:
                    cell.font = Font(
                        color="FF0000",
                        bold=True,
                    )
            else:
                cell.value = ""
                cell.alignment = CENTER

        # Auto Width
        for column in ws.columns:

            length = 0
            letter = column[0].column_letter

            for cell in column:
                if cell.value is not None:
                    length = max(length, len(str(cell.value)))

            ws.column_dimensions[letter].width = min(length + 3, 50)

        ws.freeze_panes = "A2"

    with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        wb.save(tmp.name)
        file_path = tmp.name

    return FileResponse(
        path=file_path,
        filename="inventory_movement_report.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
