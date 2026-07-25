import openpyxl
from fastapi import APIRouter, Depends
from sqlalchemy import text
from sqlalchemy.orm import Session
from tempfile import NamedTemporaryFile
from fastapi.responses import FileResponse
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from app.common.apply_payload_permissions import apply_payload_permissions
from app.core.database import get_db
from app.dependencies.auth import get_current_user
from app.reports.item_loading_report.schemas.item_loading_schema import ItemLoadingRequest
from app.reports.item_loading_report.utils.item_loading_helper import prepare_dashboard_context
from app.reports.item_loading_report.routes.item_loading_tableview import build_item_loading_query
  

router = APIRouter(tags=["Item Loading Report"], dependencies=[Depends(get_current_user)])

HEADER_FILL = PatternFill(
    start_color="903442",
    end_color="903442",
    fill_type="solid"
)

HEADER_FONT = Font(
    color="FFFFFF",
    bold=True
)

TOTAL_FONT = Font(
    color="FFFFFF",
    bold=True
)

BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")


def pretty_header(header: str) -> str:
    return header.replace("_", " ").title()

def to_float(value):
    try:
        return float(value or 0)
    except Exception:
        return 0


@router.post("/item-loading-export")
def item_loading_export(
    payload: ItemLoadingRequest,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user),
):
    payload = apply_payload_permissions(payload, db, current_user)
    ctx = prepare_dashboard_context(payload)

    query = build_item_loading_query(ctx, order_by=True)
    rows = [
        dict(row)
        for row in db.execute(text(query), ctx["params"]).mappings().all()
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Item Loading Report"

    if not rows:
        ws.append(["No Data Found"])
        ws["A1"].fill = HEADER_FILL
        ws["A1"].font = HEADER_FONT
        ws["A1"].border = BORDER
        ws["A1"].alignment = CENTER
    else:
        headers = list(rows[0].keys())

        if "salesman_id" in headers:
            headers.remove("salesman_id")

        ws.append([pretty_header(header) for header in headers])

        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = BORDER
            cell.alignment = CENTER

        for row in rows:
            ws.append([row.get(header) for header in headers])

        numeric_headers = {
            "salesman_ordered_qty",
            "received_qty",
            "diff",
            "upc",
        }

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = BORDER

                header = headers[cell.column - 1]

                if header in numeric_headers:
                    cell.alignment = RIGHT
                    cell.number_format = "#,##0.000"

                    if isinstance(cell.value, (int, float)) and cell.value < 0:
                        cell.font = Font(color="FF0000")

        total_row = ws.max_row + 1

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=total_row, column=col_idx)
            cell.fill = HEADER_FILL
            cell.font = TOTAL_FONT
            cell.border = BORDER
            cell.alignment = CENTER

            if col_idx == 1:
                cell.value = "Total"
                continue

            if header in numeric_headers and header != "upc":
                total_value = sum(to_float(row.get(header)) for row in rows)
                cell.value = total_value
                cell.alignment = RIGHT
                cell.number_format = "#,##0.000"

                if total_value < 0:
                    cell.font = Font(color="FF0000", bold=True)
            else:
                cell.value = ""

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[column_letter].width = min(max_length + 3, 50)

        ws.freeze_panes = "A2"

    with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        wb.save(tmp.name)
        file_path = tmp.name

    return FileResponse(
        path=file_path,
        filename="Item_Loading_Report.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )