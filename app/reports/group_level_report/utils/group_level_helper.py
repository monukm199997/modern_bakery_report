from io import BytesIO

from fastapi import HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy import text

from app.reports.group_level_report.schemas.group_schema import GroupLevelReportRequest
from app.reports.group_level_report.utils.group_level_sql_query import (
    DRILL_DOWN_MAP,
    FACT_AGGREGATES,
    REVENUE_METRICS,
    VOLUME_METRICS,
)


def build_group_level_filters(payload: GroupLevelReportRequest):
    """
    Filters for the listed-SKU universe (customeritem_header/detail + customer).

    The date range is NOT added here — it belongs inside the sales fact CTE, so
    the LEFT JOIN keeps every listed row. from_date/to_date are still returned
    as params because the CTE binds them.
    """
    where = ["cih.deleted_at IS NULL", "cid.deleted_at IS NULL"]
    params = {"from_date": payload.from_date, "to_date": payload.to_date}

    if payload.customer_groups_ids:
        where.append("ac.cust_group = ANY(:customer_groups_ids)")
        params["customer_groups_ids"] = payload.customer_groups_ids

    if payload.customer_groups_1_ids:
        where.append("ac.customergroup = ANY(CAST(:customer_groups_1_ids AS text[]))")
        params["customer_groups_1_ids"] = payload.customer_groups_1_ids

    if payload.customer_groups_2_ids:
        where.append("ac.customergroup2 = ANY(CAST(:customer_groups_2_ids AS text[]))")
        params["customer_groups_2_ids"] = payload.customer_groups_2_ids

    if payload.company_ids:
        where.append(
            "cih.route_id IN ("
            "SELECT route_id FROM salesman "
            "WHERE company_id = ANY(:company_ids) AND route_id IS NOT NULL"
            ")"
        )
        params["company_ids"] = payload.company_ids

    return where, params


def prepare_group_level_context(payload: GroupLevelReportRequest):

    selected = [f.lower() for f in (payload.drill_down_fields or [])]

    if not selected:
        selected = ["item"]

    select_cols = []
    group_cols = []
    order_cols = []

    for field in selected:
        if field not in DRILL_DOWN_MAP:
            raise HTTPException(
                status_code=400,
                detail="drill_down_fields may only be 'customer' or 'item'",
            )
        config = DRILL_DOWN_MAP[field]
        select_cols.append(config["select"].strip().rstrip(","))
        group_cols.append(config["group_by"])
        order_cols.append(config["group_by"])

    search_type = payload.search_type.lower()

    metric_cols = []
    if search_type in ("amount", "both"):
        metric_cols += REVENUE_METRICS
    if search_type in ("quantity", "both"):
        metric_cols += VOLUME_METRICS

    where, params = build_group_level_filters(payload)

    return {
        "select_sql": ",\n".join(select_cols + metric_cols),
        "where_sql": " AND ".join(where),
        "group_by_sql": ("GROUP BY " + ", ".join(group_cols)) if group_cols else "",
        "order_by_sql": ("ORDER BY " + ", ".join(order_cols)) if order_cols else "",
        "params": params,
    }


def _build_group_level_query(ctx, extra_select="", tail=""):

    select_sql = ctx["select_sql"]
    if extra_select:
        select_sql = f"{select_sql},\n{extra_select}"
 
    return f"""
        WITH sales_facts AS (
            SELECT
                sdh.customer_id,
                sdd.item_id,
                {FACT_AGGREGATES}
            FROM sales_documents_header sdh
            JOIN sales_documents_detail sdd ON sdd.header_id = sdh.id
            WHERE sdh.invoice_date BETWEEN :from_date AND :to_date
              AND sdh.deleted_at IS NULL
              AND sdd.deleted_at IS NULL
            GROUP BY sdh.customer_id, sdd.item_id
        )
        SELECT
            {select_sql}
        FROM customeritem_header cih
        JOIN customeritem_detail cid ON cid.header_id = cih.id
        LEFT JOIN items i ON i.id = cid.item_id
        LEFT JOIN agent_customers ac ON ac.id = cih.customer_id
        LEFT JOIN sales_facts f ON f.customer_id = cih.customer_id AND f.item_id = cid.item_id
        WHERE {ctx["where_sql"]}
        {ctx["group_by_sql"]}
        {ctx["order_by_sql"]}
        {tail}
    """
 
 
def get_group_level_rows(payload: GroupLevelReportRequest, db):
 
    ctx = prepare_group_level_context(payload)
    query = _build_group_level_query(ctx)
    rows = db.execute(text(query), ctx["params"]).mappings().all()
    return [dict(row) for row in rows]



HEADER_COLOR = "903442"
REVENUE_COLOR = "FFFFC61A" 
VOLUME_COLOR = "FF3399FF"
WHITE = "FFFFFF"
BLACK = "000000"
RED = "FF0000"


def pretty_header(name: str) -> str:
    name = name.replace("_", " ").title()

    replace_map = {
        "Gross Sales": "Gross Sales",
        "Sales Return": "Sales Return",
        "Return Percent": "Return %",
        "Net Sales": "Net Sales",
    }

    return replace_map.get(name, name)

def to_float(value):
    try:
        return float(value or 0)
    except Exception:
        return 0

def calculate_percent(return_value, gross_value):
    gross_value = to_float(gross_value)
    return_value = to_float(return_value)

    if gross_value == 0:
        return 0

    return round((return_value / gross_value) * 100, 2)


def build_group_level_excel(rows):
    """
    Turn a list of report rows into the styled Sales Report .xlsx response.

    This is the single Excel formatter shared by every sales export.
    It is data-shape driven, not report-specific:
      - any column NOT prefixed 'revenue_' / 'volume_' is rendered as a
        drill-down dimension column (this is how the `period` column from
        the time-series export renders with no extra work here).
      - revenue_* columns sit under a merged "Revenue" banner.
      - volume_*  columns sit under a merged "Volume" banner.

    Row ordering is the caller's responsibility (do the ORDER BY in SQL).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    sub_header_fill = PatternFill(
        fill_type="solid",
        start_color=HEADER_COLOR,
        end_color=HEADER_COLOR,
    )

    revenue_fill = PatternFill(
        fill_type="solid",
        start_color=REVENUE_COLOR,
        end_color=REVENUE_COLOR,
    )

    volume_fill = PatternFill(
        fill_type="solid",
        start_color=VOLUME_COLOR,
        end_color=VOLUME_COLOR,
    )

    sub_header_font = Font(bold=True, color=WHITE)
    main_header_font = Font(bold=True, color=BLACK)
    total_font = Font(bold=True, color=WHITE)
    negative_font = Font(color=RED)

    if not rows:
        ws["A1"] = "No Data Found"

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=Group_Level_Report.xlsx"
            },
        )

    all_headers = list(rows[0].keys())

    dimension_headers = [
        h for h in all_headers
        if not h.startswith("revenue_") and not h.startswith("volume_")
    ]

    revenue_headers = [
        "revenue_gross_sales",
        "revenue_sales_return",
        "revenue_return_percent",
        "revenue_net_sales",
    ]

    volume_headers = [
        "volume_gross_sales",
        "volume_sales_return",
        "volume_return_percent",
        "volume_net_sales",
    ]

    revenue_headers = [h for h in revenue_headers if h in all_headers]
    volume_headers = [h for h in volume_headers if h in all_headers]

    final_headers = dimension_headers + revenue_headers + volume_headers


    col = 1

    # Drill down columns
    for header in dimension_headers:
        ws.merge_cells(
            start_row=1,
            start_column=col,
            end_row=2,
            end_column=col,
        )

        cell = ws.cell(row=1, column=col)
        cell.value = pretty_header(header)
        cell.fill = sub_header_fill
        cell.font = sub_header_font
        cell.alignment = center
        cell.border = border

        ws.cell(row=2, column=col).fill = sub_header_fill
        ws.cell(row=2, column=col).border = border

        col += 1

    # Revenue merged header
    if revenue_headers:
        start_col = col
        end_col = col + len(revenue_headers) - 1

        ws.merge_cells(
            start_row=1,
            start_column=start_col,
            end_row=1,
            end_column=end_col,
        )

        cell = ws.cell(row=1, column=start_col)
        cell.value = "Revenue"
        cell.fill = revenue_fill
        cell.font = main_header_font
        cell.alignment = center
        cell.border = border

        for c in range(start_col, end_col + 1):
            ws.cell(row=1, column=c).fill = revenue_fill
            ws.cell(row=1, column=c).border = border

        for header in revenue_headers:
            cell = ws.cell(row=2, column=col)
            cell.value = pretty_header(header.replace("revenue_", ""))
            cell.fill = sub_header_fill
            cell.font = sub_header_font
            cell.alignment = center
            cell.border = border
            col += 1

    # Volume merged header
    if volume_headers:
        start_col = col
        end_col = col + len(volume_headers) - 1

        ws.merge_cells(
            start_row=1,
            start_column=start_col,
            end_row=1,
            end_column=end_col,
        )

        cell = ws.cell(row=1, column=start_col)
        cell.value = "Volume"
        cell.fill = volume_fill
        cell.font = main_header_font
        cell.alignment = center
        cell.border = border

        for c in range(start_col, end_col + 1):
            ws.cell(row=1, column=c).fill = volume_fill
            ws.cell(row=1, column=c).border = border

        for header in volume_headers:
            cell = ws.cell(row=2, column=col)
            cell.value = pretty_header(header.replace("volume_", ""))
            cell.fill = sub_header_fill
            cell.font = sub_header_font
            cell.alignment = center
            cell.border = border
            col += 1

    # =========================
    # Data Rows
    # =========================

    start_data_row = 3

    for row_index, row in enumerate(rows, start=start_data_row):
        for col_index, header in enumerate(final_headers, start=1):
            cell = ws.cell(row=row_index, column=col_index)
            cell.value = row.get(header, 0)
            cell.border = border

            if header.startswith("revenue_") or header.startswith("volume_"):
                cell.alignment = right

            if isinstance(cell.value, (int, float)) and cell.value < 0:
                cell.font = negative_font

    # =========================
    # Total Row
    # =========================

    total_row = start_data_row + len(rows)

    for col_index, header in enumerate(final_headers, start=1):
        cell = ws.cell(row=total_row, column=col_index)
        cell.fill = sub_header_fill
        cell.font = total_font
        cell.border = border
        cell.alignment = center

        if col_index == 1:
            cell.value = "Total"
            continue

        if header.startswith("revenue_") or header.startswith("volume_"):
            cell.alignment = right

        if header == "revenue_gross_sales":
            cell.value = sum(to_float(row.get("revenue_gross_sales")) for row in rows)

        elif header == "revenue_sales_return":
            cell.value = sum(to_float(row.get("revenue_sales_return")) for row in rows)

        elif header == "revenue_return_percent":
            total_gross = sum(to_float(row.get("revenue_gross_sales")) for row in rows)
            total_return = sum(to_float(row.get("revenue_sales_return")) for row in rows)
            cell.value = calculate_percent(total_return, total_gross)

        elif header == "revenue_net_sales":
            total_gross = sum(to_float(row.get("revenue_gross_sales")) for row in rows)
            total_return = sum(to_float(row.get("revenue_sales_return")) for row in rows)
            cell.value = total_gross - total_return

        elif header == "volume_gross_sales":
            cell.value = sum(to_float(row.get("volume_gross_sales")) for row in rows)

        elif header == "volume_sales_return":
            cell.value = sum(to_float(row.get("volume_sales_return")) for row in rows)

        elif header == "volume_return_percent":
            total_gross = sum(to_float(row.get("volume_gross_sales")) for row in rows)
            total_return = sum(to_float(row.get("volume_sales_return")) for row in rows)
            cell.value = calculate_percent(total_return, total_gross)

        elif header == "volume_net_sales":
            total_gross = sum(to_float(row.get("volume_gross_sales")) for row in rows)
            total_return = sum(to_float(row.get("volume_sales_return")) for row in rows)
            cell.value = total_gross - total_return

        else:
            cell.value = ""

        if isinstance(cell.value, (int, float)) and cell.value < 0:
            cell.font = Font(bold=True, color=RED)

    # =========================
    # Number Format
    # =========================

    for row in ws.iter_rows(min_row=start_data_row, max_row=total_row):
        for cell in row:
            header = final_headers[cell.column - 1]

            if header.startswith("revenue_"):
                cell.number_format = "#,##0.00"

            elif header.startswith("volume_"):
                cell.number_format = "#,##0"

            if header.endswith("_percent"):
                cell.number_format = "0.00"

    # =========================
    # Width / Freeze
    # =========================

    for column in ws.columns:
        max_length = 0
        col_letter = get_column_letter(column[0].column)

        for cell in column:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_length + 4

    ws.freeze_panes = "A3"

    # Dropdown filter remove
    # ws.auto_filter.ref = ws.dimensions

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=Group_Level_Report.xlsx"
        },
    )












# from io import BytesIO

# from fastapi import HTTPException
# from fastapi.responses import StreamingResponse
# from openpyxl import Workbook
# from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
# from openpyxl.utils import get_column_letter
# from sqlalchemy import text

# from app.reports.group_level_report.schemas.group_schema import GroupLevelReportRequest
# from app.reports.group_level_report.utils.group_level_sql_query import (
#     DRILL_DOWN_MAP,
#     FACT_AGGREGATES,
#     REVENUE_METRICS,
#     VOLUME_METRICS,
# )


# def build_group_level_filters(payload: GroupLevelReportRequest):
#     where = []
#     params = {}

#     # where.append("sdh.invoice_date BETWEEN :from_date AND :to_date")
#     where.append("cih.deleted_at IS NULL")
#     where.append("cid.deleted_at IS NULL")

#     params["from_date"] = payload.from_date
#     params["to_date"] = payload.to_date

#     if payload.customer_groups_ids:
#         where.append("ac.cust_group = ANY(:customer_groups_ids)")
#         params["customer_groups_ids"] = payload.customer_groups_ids

#     if payload.customer_groups_1_ids:
#         where.append("ac.customergroup = ANY(CAST(:customer_groups_1_ids AS text[]))")
#         params["customer_groups_1_ids"] = payload.customer_groups_1_ids

#     if payload.customer_groups_2_ids:
#         where.append("ac.customergroup2 = ANY(CAST(:customer_groups_2_ids AS text[]))")
#         params["customer_groups_2_ids"] = payload.customer_groups_2_ids

#     return where, params

# def prepare_group_level_context(payload: GroupLevelReportRequest):
   
#     selected = [f.lower() for f in (payload.drill_down_fields or [])]

#     select_cols = []
#     group_cols = []
#     order_cols = []

#     for field in selected:
#         if field not in DRILL_DOWN_MAP:
#             raise HTTPException(
#                 status_code=400,
#                 detail="drill_down_fields may only be 'customer' or 'item'",
#             )
#         config = DRILL_DOWN_MAP[field]
#         select_cols.append(config["select"].strip().rstrip(","))
#         group_cols.append(config["group_by"])
#         order_cols.append(config["group_by"])

#     search_type = payload.search_type.lower()

#     metric_cols = []
#     if search_type in ("amount", "both"):
#         metric_cols += REVENUE_METRICS
#     if search_type in ("quantity", "both"):
#         metric_cols += VOLUME_METRICS

#     where, params = build_group_level_filters(payload)
#     company_filter = ""
#     if payload.company_ids:
#         company_filter = "AND s.company_id = ANY(:company_ids)"
#         params["company_ids"] = payload.company_ids

#     return {
#         "select_sql": ",\n".join(select_cols + metric_cols),
#         "where_sql": " AND ".join(where),
#         "group_by_sql": ("GROUP BY " + ", ".join(group_cols)) if group_cols else "",
#         "order_by_sql": ("ORDER BY " + ", ".join(order_cols)) if order_cols else "",
#         "company_filter": company_filter,
#         "params": params,
#     }


# def get_group_level_rows(payload: GroupLevelReportRequest, db):
#     ctx = prepare_group_level_context(payload)

#     query = f"""
#         WITH sales_facts AS (
#             SELECT
#                 sdh.customer_id,
#                 sdd.item_id,
#                 {FACT_AGGREGATES}
#             FROM sales_documents_header sdh
#             JOIN sales_documents_detail sdd ON sdd.header_id = sdh.id
#             LEFT JOIN salesman s ON s.id = sdh.salesman_id
#             WHERE sdh.invoice_date BETWEEN :from_date AND :to_date
#                 {ctx["company_filter"]}
#                 AND sdh.deleted_at IS NULL
#                 AND sdd.deleted_at IS NULL
#             GROUP BY sdh.customer_id, sdd.item_id
#         )
#         SELECT
#             {ctx["select_sql"]}
#         FROM customeritem_header cih
#         JOIN customeritem_detail cid ON cid.header_id = cih.id
#         LEFT JOIN items i ON i.id = cid.item_id
#         LEFT JOIN agent_customers ac ON ac.id = cih.customer_id
#         LEFT JOIN sales_facts f ON f.customer_id = cih.customer_id AND f.item_id = cid.item_id
#         WHERE {ctx["where_sql"]}
#         {ctx["group_by_sql"]}
#         {ctx["order_by_sql"]}
#     """
#     rows = db.execute(text(query), ctx["params"]).mappings().all()
#     print(query)
#     return [dict(row) for row in rows]


# HEADER_COLOR = "903442"
# REVENUE_COLOR = "FFFFC61A" 
# VOLUME_COLOR = "FF3399FF"
# WHITE = "FFFFFF"
# BLACK = "000000"
# RED = "FF0000"


# def pretty_header(name: str) -> str:
#     name = name.replace("_", " ").title()

#     replace_map = {
#         "Gross Sales": "Gross Sales",
#         "Sales Return": "Sales Return",
#         "Return Percent": "Return %",
#         "Net Sales": "Net Sales",
#     }

#     return replace_map.get(name, name)

# def to_float(value):
#     try:
#         return float(value or 0)
#     except Exception:
#         return 0

# def calculate_percent(return_value, gross_value):
#     gross_value = to_float(gross_value)
#     return_value = to_float(return_value)

#     if gross_value == 0:
#         return 0

#     return round((return_value / gross_value) * 100, 2)


# def build_group_level_excel(rows):

#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Sales Report"

#     border = Border(
#         left=Side(style="thin"),
#         right=Side(style="thin"),
#         top=Side(style="thin"),
#         bottom=Side(style="thin"),
#     )

#     center = Alignment(horizontal="center", vertical="center")
#     right = Alignment(horizontal="right", vertical="center")

#     sub_header_fill = PatternFill(
#         fill_type="solid",
#         start_color=HEADER_COLOR,
#         end_color=HEADER_COLOR,
#     )

#     revenue_fill = PatternFill(
#         fill_type="solid",
#         start_color=REVENUE_COLOR,
#         end_color=REVENUE_COLOR,
#     )

#     volume_fill = PatternFill(
#         fill_type="solid",
#         start_color=VOLUME_COLOR,
#         end_color=VOLUME_COLOR,
#     )

#     sub_header_font = Font(bold=True, color=WHITE)
#     main_header_font = Font(bold=True, color=BLACK)
#     total_font = Font(bold=True, color=WHITE)
#     negative_font = Font(color=RED)

#     if not rows:
#         ws["A1"] = "No Data Found"

#         output = BytesIO()
#         wb.save(output)
#         output.seek(0)

#         return StreamingResponse(
#             output,
#             media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#             headers={
#                 "Content-Disposition": "attachment; filename=Group_Level_Report.xlsx"
#             },
#         )

#     all_headers = list(rows[0].keys())

#     dimension_headers = [
#         h for h in all_headers
#         if not h.startswith("revenue_") and not h.startswith("volume_")
#     ]

#     revenue_headers = [
#         "revenue_gross_sales",
#         "revenue_sales_return",
#         "revenue_return_percent",
#         "revenue_net_sales",
#     ]

#     volume_headers = [
#         "volume_gross_sales",
#         "volume_sales_return",
#         "volume_return_percent",
#         "volume_net_sales",
#     ]

#     revenue_headers = [h for h in revenue_headers if h in all_headers]
#     volume_headers = [h for h in volume_headers if h in all_headers]

#     final_headers = dimension_headers + revenue_headers + volume_headers


#     col = 1

#     # Drill down columns
#     for header in dimension_headers:
#         ws.merge_cells(
#             start_row=1,
#             start_column=col,
#             end_row=2,
#             end_column=col,
#         )

#         cell = ws.cell(row=1, column=col)
#         cell.value = pretty_header(header)
#         cell.fill = sub_header_fill
#         cell.font = sub_header_font
#         cell.alignment = center
#         cell.border = border

#         ws.cell(row=2, column=col).fill = sub_header_fill
#         ws.cell(row=2, column=col).border = border

#         col += 1

#     # Revenue merged header
#     if revenue_headers:
#         start_col = col
#         end_col = col + len(revenue_headers) - 1

#         ws.merge_cells(
#             start_row=1,
#             start_column=start_col,
#             end_row=1,
#             end_column=end_col,
#         )

#         cell = ws.cell(row=1, column=start_col)
#         cell.value = "Revenue"
#         cell.fill = revenue_fill
#         cell.font = main_header_font
#         cell.alignment = center
#         cell.border = border

#         for c in range(start_col, end_col + 1):
#             ws.cell(row=1, column=c).fill = revenue_fill
#             ws.cell(row=1, column=c).border = border

#         for header in revenue_headers:
#             cell = ws.cell(row=2, column=col)
#             cell.value = pretty_header(header.replace("revenue_", ""))
#             cell.fill = sub_header_fill
#             cell.font = sub_header_font
#             cell.alignment = center
#             cell.border = border
#             col += 1

#     # Volume merged header
#     if volume_headers:
#         start_col = col
#         end_col = col + len(volume_headers) - 1

#         ws.merge_cells(
#             start_row=1,
#             start_column=start_col,
#             end_row=1,
#             end_column=end_col,
#         )

#         cell = ws.cell(row=1, column=start_col)
#         cell.value = "Volume"
#         cell.fill = volume_fill
#         cell.font = main_header_font
#         cell.alignment = center
#         cell.border = border

#         for c in range(start_col, end_col + 1):
#             ws.cell(row=1, column=c).fill = volume_fill
#             ws.cell(row=1, column=c).border = border

#         for header in volume_headers:
#             cell = ws.cell(row=2, column=col)
#             cell.value = pretty_header(header.replace("volume_", ""))
#             cell.fill = sub_header_fill
#             cell.font = sub_header_font
#             cell.alignment = center
#             cell.border = border
#             col += 1

#     # =========================
#     # Data Rows
#     # =========================

#     start_data_row = 3

#     for row_index, row in enumerate(rows, start=start_data_row):
#         for col_index, header in enumerate(final_headers, start=1):
#             cell = ws.cell(row=row_index, column=col_index)
#             cell.value = row.get(header, 0)
#             cell.border = border

#             if header.startswith("revenue_") or header.startswith("volume_"):
#                 cell.alignment = right

#             if isinstance(cell.value, (int, float)) and cell.value < 0:
#                 cell.font = negative_font

#     # =========================
#     # Total Row
#     # =========================

#     total_row = start_data_row + len(rows)

#     for col_index, header in enumerate(final_headers, start=1):
#         cell = ws.cell(row=total_row, column=col_index)
#         cell.fill = sub_header_fill
#         cell.font = total_font
#         cell.border = border
#         cell.alignment = center

#         if col_index == 1:
#             cell.value = "Total"
#             continue

#         if header.startswith("revenue_") or header.startswith("volume_"):
#             cell.alignment = right

#         if header == "revenue_gross_sales":
#             cell.value = sum(to_float(row.get("revenue_gross_sales")) for row in rows)

#         elif header == "revenue_sales_return":
#             cell.value = sum(to_float(row.get("revenue_sales_return")) for row in rows)

#         elif header == "revenue_return_percent":
#             total_gross = sum(to_float(row.get("revenue_gross_sales")) for row in rows)
#             total_return = sum(to_float(row.get("revenue_sales_return")) for row in rows)
#             cell.value = calculate_percent(total_return, total_gross)

#         elif header == "revenue_net_sales":
#             total_gross = sum(to_float(row.get("revenue_gross_sales")) for row in rows)
#             total_return = sum(to_float(row.get("revenue_sales_return")) for row in rows)
#             cell.value = total_gross - total_return

#         elif header == "volume_gross_sales":
#             cell.value = sum(to_float(row.get("volume_gross_sales")) for row in rows)

#         elif header == "volume_sales_return":
#             cell.value = sum(to_float(row.get("volume_sales_return")) for row in rows)

#         elif header == "volume_return_percent":
#             total_gross = sum(to_float(row.get("volume_gross_sales")) for row in rows)
#             total_return = sum(to_float(row.get("volume_sales_return")) for row in rows)
#             cell.value = calculate_percent(total_return, total_gross)

#         elif header == "volume_net_sales":
#             total_gross = sum(to_float(row.get("volume_gross_sales")) for row in rows)
#             total_return = sum(to_float(row.get("volume_sales_return")) for row in rows)
#             cell.value = total_gross - total_return

#         else:
#             cell.value = ""

#         if isinstance(cell.value, (int, float)) and cell.value < 0:
#             cell.font = Font(bold=True, color=RED)

#     # =========================
#     # Number Format
#     # =========================

#     for row in ws.iter_rows(min_row=start_data_row, max_row=total_row):
#         for cell in row:
#             header = final_headers[cell.column - 1]

#             if header.startswith("revenue_"):
#                 cell.number_format = "#,##0.00"

#             elif header.startswith("volume_"):
#                 cell.number_format = "#,##0"

#             if header.endswith("_percent"):
#                 cell.number_format = "0.00"

#     # =========================
#     # Width / Freeze
#     # =========================

#     for column in ws.columns:
#         max_length = 0
#         col_letter = get_column_letter(column[0].column)

#         for cell in column:
#             if cell.value is not None:
#                 max_length = max(max_length, len(str(cell.value)))

#         ws.column_dimensions[col_letter].width = max_length + 4

#     ws.freeze_panes = "A3"

#     # Dropdown filter remove
#     # ws.auto_filter.ref = ws.dimensions

#     output = BytesIO()
#     wb.save(output)
#     output.seek(0)

#     return StreamingResponse(
#         output,
#         media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#         headers={
#             "Content-Disposition": "attachment; filename=Group_Level_Report.xlsx"
#         },
#     )