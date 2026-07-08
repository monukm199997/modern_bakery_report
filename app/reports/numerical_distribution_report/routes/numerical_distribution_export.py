from io import BytesIO

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse

from sqlalchemy import text
from sqlalchemy.orm import Session

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from app.core.database import get_db
from app.dependencies.auth import get_current_user

from app.reports.numerical_distribution_report.schemas.numerical_distribution_schema import (
    NumericalDistributionRequest,
)

from app.reports.numerical_distribution_report.utils.numerical_distribution_helper import (
    prepare_numerical_distribution_context,
)

router = APIRouter(tags=["Numerical Distribution Report"])


@router.post("/export")
def export_numerical_distribution(
    payload: NumericalDistributionRequest,
    db: Session = Depends(get_db)
):
    ctx = prepare_numerical_distribution_context(payload)
    sql = f"""
        SELECT
            {ctx["select"]}
        {ctx["from"]}
        WHERE
            {ctx["where"]}
        GROUP BY
            {ctx["group_by"]}
        ORDER BY
            {ctx["order_by"]}
    """
    rows = db.execute(text(sql), ctx["params"],).mappings().all() 

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Numerical Distribution"

    HEADER_FILL = PatternFill(
        fill_type="solid",
        fgColor="903442"
    )

    HEADER_FONT = Font(
        bold=True,
        color="FFFFFF",
        size=11,
    )

    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    CENTER = Alignment(
        horizontal="center",
        vertical="center",
    )

    if not rows:
        sheet.cell(row=1, column=1).value = "No Data Found"
        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition":
                    "attachment; filename=Numerical_Distribution_Report.xlsx"
            },
        )


    headers = list(rows[0].keys())
    for col_num, header in enumerate(headers, start=1):

        cell = sheet.cell(row=1, column=col_num)

        cell.value = header.replace("_", " ").title()

        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = CENTER



    data_alignment = Alignment(
        vertical="center",
        horizontal="left",
    )

    number_alignment = Alignment(
        vertical="center",
        horizontal="right",
    )

    for row_num, row in enumerate(rows, start=2):

        for col_num, value in enumerate(row.values(), start=1):
            cell = sheet.cell(
                row=row_num,
                column=col_num,
            )
            cell.value = value
            cell.border = THIN_BORDER
            if isinstance(value, (int, float)):
                cell.alignment = number_alignment
            else:
                cell.alignment = data_alignment

    sheet.freeze_panes = "A2"

    for column_cells in sheet.columns:
        max_length = 0
        column = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            try:
                if cell.value is not None:
                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )
            except Exception:
                pass

        adjusted_width = min(max(max_length + 4, 12), 50)

        sheet.column_dimensions[column].width = adjusted_width
    # sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = True
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = (
        "Numerical_Distribution_report.xlsx"
    )
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"'
        },
    )