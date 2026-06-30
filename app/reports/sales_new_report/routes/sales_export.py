from fastapi import APIRouter, Depends
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from app.core.database import get_db
from app.dependencies.auth import get_current_user
from app.reports.sales_new_report.schemas.sales_schema import SalesReportRequest
from app.reports.sales_new_report.routes.sales_tableview import get_sales_report_table

router = APIRouter(tags=["Sales New Report"], dependencies=[Depends(get_current_user)])


HEADER_COLOR = "903442"
REVENUE_COLOR = "FFFFC61A" 
VOLUME_COLOR = "FF3399FF"
WHITE = "FFFFFF"
BLACK = "000000"
RED = "FF0000"


def pretty_header(name: str) -> str:
    name = name.replace("_", " ").title()

    replace_map = {
        "Gross Sales": "Gross Sales",
        "Sales Return": "Sales Return",
        "Return Percent": "Return %",
        "Net Sales": "Net Sales",
    }

    return replace_map.get(name, name)

def to_float(value):
    try:
        return float(value or 0)
    except Exception:
        return 0

def calculate_percent(return_value, gross_value):
    gross_value = to_float(gross_value)
    return_value = to_float(return_value)

    if gross_value == 0:
        return 0

    return round((return_value / gross_value) * 100, 2)

def export_sales_report(payload, db):
    result = get_sales_report_table(payload, db)
    rows = result["data"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    sub_header_fill = PatternFill(
        fill_type="solid",
        start_color=HEADER_COLOR,
        end_color=HEADER_COLOR,
    )

    revenue_fill = PatternFill(
        fill_type="solid",
        start_color=REVENUE_COLOR,
        end_color=REVENUE_COLOR,
    )

    volume_fill = PatternFill(
        fill_type="solid",
        start_color=VOLUME_COLOR,
        end_color=VOLUME_COLOR,
    )

    sub_header_font = Font(bold=True, color=WHITE)
    main_header_font = Font(bold=True, color=BLACK)
    total_font = Font(bold=True, color=WHITE)
    negative_font = Font(color=RED)

    if not rows:
        ws["A1"] = "No Data Found"

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=Sales_Report.xlsx"
            },
        )

    all_headers = list(rows[0].keys())

    dimension_headers = [
        h for h in all_headers
        if not h.startswith("revenue_") and not h.startswith("volume_")
    ]

    revenue_headers = [
        "revenue_gross_sales",
        "revenue_sales_return",
        "revenue_return_percent",
        "revenue_net_sales",
    ]

    volume_headers = [
        "volume_gross_sales",
        "volume_sales_return",
        "volume_return_percent",
        "volume_net_sales",
    ]

    revenue_headers = [h for h in revenue_headers if h in all_headers]
    volume_headers = [h for h in volume_headers if h in all_headers]

    final_headers = dimension_headers + revenue_headers + volume_headers


    col = 1

    # Drill down columns
    for header in dimension_headers:
        ws.merge_cells(
            start_row=1,
            start_column=col,
            end_row=2,
            end_column=col,
        )

        cell = ws.cell(row=1, column=col)
        cell.value = pretty_header(header)
        cell.fill = sub_header_fill
        cell.font = sub_header_font
        cell.alignment = center
        cell.border = border

        ws.cell(row=2, column=col).fill = sub_header_fill
        ws.cell(row=2, column=col).border = border

        col += 1

    # Revenue merged header
    if revenue_headers:
        start_col = col
        end_col = col + len(revenue_headers) - 1

        ws.merge_cells(
            start_row=1,
            start_column=start_col,
            end_row=1,
            end_column=end_col,
        )

        cell = ws.cell(row=1, column=start_col)
        cell.value = "Revenue"
        cell.fill = revenue_fill
        cell.font = main_header_font
        cell.alignment = center
        cell.border = border

        for c in range(start_col, end_col + 1):
            ws.cell(row=1, column=c).fill = revenue_fill
            ws.cell(row=1, column=c).border = border

        for header in revenue_headers:
            cell = ws.cell(row=2, column=col)
            cell.value = pretty_header(header.replace("revenue_", ""))
            cell.fill = sub_header_fill
            cell.font = sub_header_font
            cell.alignment = center
            cell.border = border
            col += 1

    # Volume merged header
    if volume_headers:
        start_col = col
        end_col = col + len(volume_headers) - 1

        ws.merge_cells(
            start_row=1,
            start_column=start_col,
            end_row=1,
            end_column=end_col,
        )

        cell = ws.cell(row=1, column=start_col)
        cell.value = "Volume"
        cell.fill = volume_fill
        cell.font = main_header_font
        cell.alignment = center
        cell.border = border

        for c in range(start_col, end_col + 1):
            ws.cell(row=1, column=c).fill = volume_fill
            ws.cell(row=1, column=c).border = border

        for header in volume_headers:
            cell = ws.cell(row=2, column=col)
            cell.value = pretty_header(header.replace("volume_", ""))
            cell.fill = sub_header_fill
            cell.font = sub_header_font
            cell.alignment = center
            cell.border = border
            col += 1

    # =========================
    # Data Rows
    # =========================

    start_data_row = 3

    for row_index, row in enumerate(rows, start=start_data_row):
        for col_index, header in enumerate(final_headers, start=1):
            cell = ws.cell(row=row_index, column=col_index)
            cell.value = row.get(header, 0)
            cell.border = border

            if header.startswith("revenue_") or header.startswith("volume_"):
                cell.alignment = right

            if isinstance(cell.value, (int, float)) and cell.value < 0:
                cell.font = negative_font

    # =========================
    # Total Row
    # =========================

    total_row = start_data_row + len(rows)

    for col_index, header in enumerate(final_headers, start=1):
        cell = ws.cell(row=total_row, column=col_index)
        cell.fill = sub_header_fill
        cell.font = total_font
        cell.border = border
        cell.alignment = center

        if col_index == 1:
            cell.value = "Total"
            continue

        if header.startswith("revenue_") or header.startswith("volume_"):
            cell.alignment = right

        if header == "revenue_gross_sales":
            cell.value = sum(to_float(row.get("revenue_gross_sales")) for row in rows)

        elif header == "revenue_sales_return":
            cell.value = sum(to_float(row.get("revenue_sales_return")) for row in rows)

        elif header == "revenue_return_percent":
            total_gross = sum(to_float(row.get("revenue_gross_sales")) for row in rows)
            total_return = sum(to_float(row.get("revenue_sales_return")) for row in rows)
            cell.value = calculate_percent(total_return, total_gross)

        elif header == "revenue_net_sales":
            total_gross = sum(to_float(row.get("revenue_gross_sales")) for row in rows)
            total_return = sum(to_float(row.get("revenue_sales_return")) for row in rows)
            cell.value = total_gross - total_return

        elif header == "volume_gross_sales":
            cell.value = sum(to_float(row.get("volume_gross_sales")) for row in rows)

        elif header == "volume_sales_return":
            cell.value = sum(to_float(row.get("volume_sales_return")) for row in rows)

        elif header == "volume_return_percent":
            total_gross = sum(to_float(row.get("volume_gross_sales")) for row in rows)
            total_return = sum(to_float(row.get("volume_sales_return")) for row in rows)
            cell.value = calculate_percent(total_return, total_gross)

        elif header == "volume_net_sales":
            total_gross = sum(to_float(row.get("volume_gross_sales")) for row in rows)
            total_return = sum(to_float(row.get("volume_sales_return")) for row in rows)
            cell.value = total_gross - total_return

        else:
            cell.value = ""

        if isinstance(cell.value, (int, float)) and cell.value < 0:
            cell.font = Font(bold=True, color=RED)

    # =========================
    # Number Format
    # =========================

    for row in ws.iter_rows(min_row=start_data_row, max_row=total_row):
        for cell in row:
            header = final_headers[cell.column - 1]

            if header.startswith("revenue_"):
                cell.number_format = "#,##0.00"

            elif header.startswith("volume_"):
                cell.number_format = "#,##0"

            if header.endswith("_percent"):
                cell.number_format = "0.00"

    # =========================
    # Width / Freeze
    # =========================

    for column in ws.columns:
        max_length = 0
        col_letter = get_column_letter(column[0].column)

        for cell in column:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_length + 4

    ws.freeze_panes = "A3"

    # Dropdown filter remove
    # ws.auto_filter.ref = ws.dimensions

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=Sales_Report.xlsx"
        },
    )

@router.post("/sales-new-export")
def export_report(
    payload: SalesReportRequest,
    db: Session = Depends(get_db),
):
    return export_sales_report(payload, db)