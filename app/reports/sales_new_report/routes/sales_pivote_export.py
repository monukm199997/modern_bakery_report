from datetime import date
from io import BytesIO

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy import text
from sqlalchemy.orm import Session
from app.common.apply_payload_permissions import apply_payload_permissions
from app.core.database import get_db
from app.dependencies.auth import get_current_user
from app.reports.sales_new_report.schemas.sales_schema import SalesPivotExportRequest
from app.reports.sales_new_report.utils.sales_helper import prepare_sales_pivot_context

router = APIRouter(tags=["Sales New Report"])

HEADER_COLOR = "903442"
REVENUE_COLOR = "FFFFC61A"
VOLUME_COLOR = "FF3399FF"
WHITE = "FFFFFF"
BLACK = "000000"
RED = "FF0000"

def get_sales_pivot_rows(payload: SalesPivotExportRequest, db: Session):
    ctx = prepare_sales_pivot_context(payload)

    query = f"""
        SELECT
            {ctx["select_sql"]}
        FROM sales_documents_header sdh
        JOIN sales_documents_detail sdd ON sdd.header_id = sdh.id
        LEFT JOIN salesman sm ON sm.id = sdh.salesman_id
        LEFT JOIN users sup ON sup.id = sm.superwiser_id AND sup.role = 108
        LEFT JOIN items i ON i.id = sdd.item_id
        LEFT JOIN agent_customers ac ON ac.id = sdh.customer_id
        LEFT JOIN outlet_channel oc ON oc.id = ac.outlet_channel_id
        LEFT JOIN tbl_route rt ON rt.id = sdh.route_id
        {ctx["extra_join_sql"]}
        WHERE {ctx["where_sql"]}
        {ctx["group_by_sql"]}
        {ctx["order_by_sql"]}
    """
    rows = db.execute(text(query), ctx["params"]).mappings().all()
    return [dict(row) for row in rows]


def _parse_date(value):
    if isinstance(value, date):
        return value
    try:
        return date.fromisoformat(str(value)[:10])
    except Exception:
        return None


def _period_label(period_start: date, period: str) -> str:
    if period == "day":
        return period_start.strftime("%Y-%m-%d")

    if period == "month":
        return period_start.strftime("%b %Y")

    if period == "year":
        return period_start.strftime("%Y")

    return str(period_start)

REVENUE_BLOCK = ("Revenue", REVENUE_COLOR, [
    ("revenue_gross_sales",   "Gross Sales",  "#,##0.00"),
    ("revenue_sales_return",  "Sales Return", "#,##0.00"),
    ("revenue_return_percent", "Return %",    "0.00"),
    ("revenue_net_sales",     "Net Sales",    "#,##0.00"),
])
VOLUME_BLOCK = ("Volume", VOLUME_COLOR, [
    ("volume_gross_sales",    "Gross Sales",  "#,##0.000"),
    ("volume_sales_return",   "Sales Return", "#,##0.000"),
    ("volume_return_percent", "Return %",     "0.00"),
    ("volume_net_sales",      "Net Sales",    "#,##0.000"),
])
 
METRICS_PER_PERIOD = 4
 
 
def _blocks_for(search_type):
    if search_type == "amount":
        return [REVENUE_BLOCK]
    if search_type == "quantity":
        return [VOLUME_BLOCK]
    return [REVENUE_BLOCK, VOLUME_BLOCK]

def build_sales_pivot_excel(rows, period, search_type):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"
 
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    header_fill = PatternFill(fill_type="solid", start_color=HEADER_COLOR, end_color=HEADER_COLOR)
    header_font = Font(bold=True, color=WHITE)
    banner_font = Font(bold=True, color=BLACK)
 
    def finish():
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=Sales_Report.xlsx"},
        )
 
    if not rows:
        ws["A1"] = "No Data Found"
        return finish()
 
    blocks = _blocks_for(search_type)
 
    reserved = {"period_start"}
    entity_cols = [
        k for k in rows[0].keys()
        if k not in reserved and not k.startswith("revenue_") and not k.startswith("volume_")
    ]
 
    periods = sorted({_parse_date(r["period_start"]) for r in rows if r["period_start"] is not None})
    period_labels = [_period_label(p, period) for p in periods]
 
    row_order = []
    row_meta = {}
    for r in rows:
        key = tuple(r[c] for c in entity_cols)
        if key not in row_meta:
            row_order.append(key)
            row_meta[key] = [r[c] for c in entity_cols]
 
    if entity_cols:
        leading_headers = list(entity_cols)          # raw names, like the reference
    else:
        leading_headers = ["summary"]
        row_meta = {k: ["All"] for k in row_meta}
    n_lead = len(leading_headers)
 
    # matrix: {row_key -> {period_start -> {metric_key -> value}}}
    metric_keys = [m[0] for blk in blocks for m in blk[2]]
    matrix = {key: {} for key in row_order}
    for r in rows:
        key = tuple(r[c] for c in entity_cols)
        p = _parse_date(r["period_start"])
        cell_vals = {}
        for mk in metric_keys:
            try:
                cell_vals[mk] = float(r.get(mk) or 0)
            except (TypeError, ValueError):
                cell_vals[mk] = 0.0
        matrix[key][p] = cell_vals
 
    block_width = len(periods) * METRICS_PER_PERIOD  # columns spanned by one block
 
    # =========================================================
    # Header row 1: drill-down headers (merged down 3 rows) + block banners
    # =========================================================
    for c in range(1, n_lead + 1):
        ws.merge_cells(start_row=1, start_column=c, end_row=3, end_column=c)
        cell = ws.cell(row=1, column=c)
        cell.value = leading_headers[c - 1]
        cell.font = header_font
        cell.alignment = center

        for rr in (1, 2, 3):
            edge = ws.cell(row=rr, column=c)
            edge.fill = header_fill
            edge.border = border
 
    col = n_lead + 1
    for title, fill_color, _sub in blocks:
        start_col = col
        end_col = col + block_width - 1
        banner_fill = PatternFill(fill_type="solid", start_color=fill_color, end_color=fill_color)
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        cell = ws.cell(row=1, column=start_col)
        cell.value = title
        cell.fill = banner_fill
        cell.font = banner_font
        cell.alignment = center
        cell.border = border
        for cc in range(start_col, end_col + 1):
            ws.cell(row=1, column=cc).fill = banner_fill
            ws.cell(row=1, column=cc).border = border
        col = end_col + 1
 
    # =========================================================
    # Header row 2: date label per period (merged across the 4 metric cols)
    # =========================================================
    col = n_lead + 1
    for _title, _fill, _sub in blocks:
        for label in period_labels:
            start_col = col
            end_col = col + METRICS_PER_PERIOD - 1
            ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
            cell = ws.cell(row=2, column=start_col)
            cell.value = label
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
            for cc in range(start_col, end_col + 1):
                ws.cell(row=2, column=cc).fill = header_fill
                ws.cell(row=2, column=cc).border = border
            col = end_col + 1
 
    # =========================================================
    # Header row 3: metric sub-headers, repeated per period per block
    # =========================================================
    col = n_lead + 1
    for _title, _fill, submetrics in blocks:
        for _p in periods:
            for _key, sub_label, _fmt in submetrics:
                cell = ws.cell(row=3, column=col)
                cell.value = sub_label
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center
                cell.border = border
                col += 1
 
    # =========================================================
    # Data rows (start row 4)
    # =========================================================
    r_idx = 4
    for key in row_order:
        for c, value in enumerate(row_meta[key], start=1):
            cell = ws.cell(row=r_idx, column=c)
            cell.value = value
            cell.border = border
 
        col = n_lead + 1
        for _title, _fill, submetrics in blocks:
            for p in periods:
                cell_vals = matrix[key].get(p, {})
                for metric_key, _label, num_fmt in submetrics:
                    v = cell_vals.get(metric_key, 0.0)
                    cell = ws.cell(row=r_idx, column=col)
                    cell.value = v
                    cell.alignment = right
                    cell.border = border
                    cell.number_format = num_fmt
                    if v < 0:
                        cell.font = Font(color=RED)
                    col += 1
        r_idx += 1
 
    # =========================================================
    # Widths + freeze (freeze the 3 header rows only, so all columns scroll)
    # =========================================================
    for column in ws.columns:
        max_length = 0
        col_letter = get_column_letter(column[0].column)
        for cell in column:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 3
 
    ws.freeze_panes = "A4"
 
    return finish()
 

@router.post("/sales-pivot-export")
def export_pivot_report(
    payload: SalesPivotExportRequest,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    payload = apply_payload_permissions(payload, db, current_user)
    rows = get_sales_pivot_rows(payload, db)
    return build_sales_pivot_excel(
        rows,
        period=payload.period.lower(),
        search_type=payload.search_type.lower(),
    )