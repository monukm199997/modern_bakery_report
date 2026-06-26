
from io import BytesIO
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

from app.core.database import get_db
from app.dependencies.auth import get_current_user
from app.reports.team_master_report.schemas.team_master_schema import TeamMasterRequest
from app.reports.team_master_report.utils.team_master_helper import prepare_dashboard_context

router = APIRouter(tags=["Team Master Report"], dependencies=[Depends(get_current_user)])


HEADER_FILL = PatternFill(
    start_color="903442",
    end_color="903442",
    fill_type="solid"
)

HEADER_FONT = Font(
    bold=True,
    color="FFFFFF"
)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

CENTER = Alignment(horizontal="center", vertical="center")


@router.post("/team-master-export")
def export_team_master(
    payload: TeamMasterRequest,
    db: Session = Depends(get_db)
):
    ctx = prepare_dashboard_context(payload)

    query = f"""
        SELECT
            rt.route_code,
            rt.route_name,
            s.osa_code AS salesman_code,
            s.name AS salesman_name,
            s.dateof_join,
            r.region_name
        FROM salesman s
        LEFT JOIN tbl_route rt
            ON s.route_id = rt.id
        LEFT JOIN tbl_region r
            ON rt.region_id = r.id
        WHERE {ctx['where_sql']}
        ORDER BY rt.route_code, s.osa_code
    """

    rows = db.execute(text(query), ctx["params"]).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Team Master"

    headers = [
        "Route Code",
        "Route Name",
        "Salesman Code",
        "Salesman Name",
        "Date of Join",
        "Region"
    ]

    ws.append(headers)

    # Header Style
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = CENTER

    # Data
    for row in rows:
        ws.append([
            row.route_code,
            row.route_name,
            row.salesman_code,
            row.salesman_name,
            row.dateof_join.strftime("%d-%m-%Y") if row.dateof_join else "",
            row.region_name
        ])

    # Apply border to all cells
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = THIN_BORDER

    # Auto width
    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(length + 4, 40)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="Team_Master_Report.xlsx"'
        },
    )