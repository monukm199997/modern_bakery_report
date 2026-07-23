import io
from datetime import datetime

import openpyxl
from openpyxl.styles import PatternFill, Font
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text
from app.common.apply_payload_permissions import apply_payload_permissions
from app.reports.vehicles_report.schemas.vehicles_schema import VehiclesRequest
from app.core.database import get_db
from app.dependencies.auth import get_current_user
from app.reports.vehicles_report.utils.vehicles_helper import prepare_dashboard_context

router = APIRouter(tags=["vehicles_report"], dependencies=[Depends(get_current_user)])

DB_COLUMNS = ["trip_date", "trip_code", "vehicle_code", "vehicle_no_plat", "vehicle_chesis_no",  "vehicle_type", "region_code", "region", "route_code", "route", "salesman_code", "salesman", "superwiser", "start_odometer", "end_odometer", "distance_traveled"]
HEADERS = ["Trip Date", "Trip Code", "Vehicle Code", "Vehicle Number Plat", "Vehicle Chesis Number", "Vehicle Type", "Region Code", "Region", "Route Code", "Route", "Sales Team Code", "Sales Team", "Superwiser", "Start Odometer", "End Odometer", "Total Distance"]
HEADER_FILL = PatternFill(start_color="FF993442", end_color="FF993442", fill_type="solid")
HEADER_FONT = Font(color="FFFFFFFF", bold=True)

@router.post("/vehicle-export")
def vehicle_export(
    payload: VehiclesRequest, 
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
    ):
    payload = apply_payload_permissions(payload, db, current_user)
    ctx = prepare_dashboard_context(payload)
    query = f"""
        SELECT 
            r.region_code,
            r.region_name AS region,
            rt.route_code,
            rt.route_name AS route,
            s.name AS salesman,
            s.osa_code AS salesman_code,
            sup.name AS superwiser,
            tt.trip_date, 
            tt.trip_code, 
            tv.number_plat AS vehicle_no_plat,
            tv.vehicle_chesis_no,
            tv.vehicle_code,
            tv.vehicle_type,
            tt.start_odometer, 
            tt.end_odometer, 
            tt.distance_traveled
        FROM tbl_trip tt
        LEFT JOIN tbl_vehicle tv ON tv.id = tt.vehicle_id
        LEFT JOIN tbl_route rt ON rt.vehicle_id = tt.vehicle_id
        LEFT JOIN salesman s ON s.route_id = rt.id
        LEFT JOIN users sup ON sup.id = s.superwiser_id AND sup.role = 108
        LEFT JOIN tbl_region r ON r.id = rt.region_id
        WHERE {ctx['where_sql']}
        ORDER BY trip_date DESC
    """
    rows = db.execute(text(query), ctx["params"]).mappings().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trips"
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for r in rows:
        ws.append([r[c] for c in DB_COLUMNS])

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"vehicle_trips_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )