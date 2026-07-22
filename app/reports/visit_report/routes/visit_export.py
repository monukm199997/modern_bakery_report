from fastapi import APIRouter, Depends
from fastapi.responses import FileResponse
from sqlalchemy import text
from sqlalchemy.orm import Session
from tempfile import NamedTemporaryFile
import openpyxl

from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from app.common.apply_payload_permissions import apply_payload_permissions
from app.core.database import get_db
from app.dependencies.auth import get_current_user
from app.reports.visit_report.schemas.visit_schema import VisitPlanRequest
from app.reports.visit_report.utils.visit_helper import prepare_dashboard_context
from app.reports.visit_report.utils.visit_sql_query import BASE_SQL_JOIN, SELECT_SQL

router = APIRouter(tags=["Visit Report"], dependencies=[Depends(get_current_user)])

HEADER_FILL = PatternFill(
    start_color="993442",
    end_color="993442",
    fill_type="solid"
)

HEADER_FONT = Font(
    color="FFFFFF",
    bold=True
)

BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

CENTER = Alignment(horizontal="center", vertical="center")


@router.post("/visit-export")
def visit_export(
    payload: VisitPlanRequest,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    payload = apply_payload_permissions(payload, db, current_user)
    ctx = prepare_dashboard_context(payload)

    query = f"""
        SELECT
            {SELECT_SQL}
        {BASE_SQL_JOIN}
        WHERE {ctx['where_sql']}
        ORDER BY date 
    """

    rows = db.execute(text(query), ctx["params"]).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Visit Report"

    headers = [
        "Date",
        "Customer Code",
        "Customer Name",
        "Customer Contact",
        "Route Code",
        "Route",
        "Sales Team Code",
        "Sales Team",
        "Superwiser",
        "Time In",
        "Time Out",
        "Spend Time",
        "Idle Time",
        "Login Time",
        "Customer Latitude",
        "Customer Longitude",
    ]

    ws.append(headers)

    # Header Styling
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = CENTER

    # Data Rows
    for row in rows:
        ws.append([
            row.date,
            row.customer_code,
            row.customer_name,
            row.customer_contact,
            row.route_code,
            row.route_name,
            row.salesman_code,
            row.salesman_name,
            row.superwiser,
            row.visit_start_time,
            row.visit_end_time,
            row.time_spent,
            row.idle_time,
            row.login_time,
            row.customer_latitude,
            row.customer_longitude,
        ])

    # Apply Border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = BORDER

    # Auto Width
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass

        ws.column_dimensions[column_letter].width = min(max_length + 3, 50)

    with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        wb.save(tmp.name)
        file_path = tmp.name

    return FileResponse(
        path=file_path,
        filename="Visit_Report.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )